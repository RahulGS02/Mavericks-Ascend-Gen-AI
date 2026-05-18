"""
HR Analytics API Endpoints
Provides actionable insights for HR decision-making
"""
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, case, extract
from datetime import datetime, timedelta, date, timezone
from typing import Dict, List, Any, Optional
import logging
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart

from app.database import get_db
from app.models import *
from app.services.auth import get_hr_user

router = APIRouter()
logger = logging.getLogger(__name__)


def utc_now():
    """Get current UTC time as timezone-aware datetime"""
    return datetime.now(timezone.utc)


@router.get("/overview")
async def get_analytics_overview(
    days: int = 30,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_hr_user)
):
    """
    Get comprehensive HR analytics overview
    
    Provides 10+ critical insights for HR decision-making:
    1. Training Effectiveness - Are batches successful?
    2. Deployment Pipeline Health - Time to deploy?
    3. Skill Gap Analysis - What skills are missing?
    4. At-Risk Mavericks - Who needs intervention?
    5. Batch Performance Comparison
    6. Assessment Trends
    7. Trainer Effectiveness
    8. Resource Utilization
    9. Profile Review Backlog
    10. Deployment Success Rate
    """
    
    try:
        cutoff_date = utc_now() - timedelta(days=days)
        
        # ========================================
        # 1. TRAINING EFFECTIVENESS
        # ========================================
        total_assessment_attempts = db.query(func.count(AssessmentAttempt.id)).filter(
            AssessmentAttempt.evaluated_at >= cutoff_date
        ).scalar() or 0
        
        passed_attempts = db.query(func.count(AssessmentAttempt.id)).filter(
            and_(
                AssessmentAttempt.evaluated_at >= cutoff_date,
                AssessmentAttempt.passed == True
            )
        ).scalar() or 0
        
        training_success_rate = round((passed_attempts / total_assessment_attempts * 100), 1) if total_assessment_attempts > 0 else 0
        
        # ========================================
        # 2. DEPLOYMENT PIPELINE HEALTH
        # ========================================
        # Average time from approval to deployment
        deployed_mavericks = db.query(Maverick).filter(
            Maverick.deployment_status == DeploymentStatus.DEPLOYED
        ).all()
        
        total_time_to_deploy = 0
        count_with_times = 0
        
        for mav in deployed_mavericks:
            if mav.created_at:
                # Use created_at as proxy for approval date
                days_to_deploy = (utc_now() - mav.created_at).days
                total_time_to_deploy += days_to_deploy
                count_with_times += 1
        
        avg_time_to_deploy = round(total_time_to_deploy / count_with_times) if count_with_times > 0 else 0
        
        # ========================================
        # 3. SKILL GAP ANALYSIS
        # ========================================
        # Get all skills from mavericks
        all_mavericks = db.query(Maverick).filter(
            Maverick.profile_status == ProfileStatus.APPROVED
        ).all()
        
        skill_frequency = {}
        low_proficiency_skills = {}
        
        for mav in all_mavericks:
            # Count skill frequency
            all_skills = (mav.skills or []) + (mav.ai_extracted_skills or [])
            for skill in all_skills:
                skill_frequency[skill] = skill_frequency.get(skill, 0) + 1
        
        # Get skills with low proficiency from maverick_skills table
        low_prof_skills = db.query(
            MaverickSkill.skill_name,
            func.avg(MaverickSkill.proficiency_score).label('avg_score'),
            func.count(MaverickSkill.id).label('count')
        ).group_by(MaverickSkill.skill_name).having(
            func.avg(MaverickSkill.proficiency_score) < 60
        ).all()
        
        skill_gaps = [
            {"skill": skill, "avg_proficiency": round(float(avg), 1), "maverick_count": count}
            for skill, avg, count in low_prof_skills[:10]
        ]
        
        # ========================================
        # 4. AT-RISK MAVERICKS
        # ========================================
        thirty_days_ago = utc_now() - timedelta(days=30)
        
        # Mavericks with multiple failed attempts
        at_risk_mavericks = db.query(
            AssessmentAttempt.maverick_id,
            Maverick.name,
            func.count(AssessmentAttempt.id).label('failed_count')
        ).join(Maverick).filter(
            and_(
                AssessmentAttempt.passed == False,
                AssessmentAttempt.evaluated_at >= thirty_days_ago
            )
        ).group_by(AssessmentAttempt.maverick_id, Maverick.name).having(
            func.count(AssessmentAttempt.id) >= 2
        ).all()
        
        at_risk_list = [
            {"maverick_id": str(mav_id), "name": name, "failed_attempts": count}
            for mav_id, name, count in at_risk_mavericks[:10]
        ]
        
        # ========================================
        # 5. BATCH PERFORMANCE COMPARISON
        # ========================================
        batch_performance = db.query(
            Batch.id,
            Batch.name,
            func.count(AssessmentAttempt.id).label('total_attempts'),
            func.sum(case((AssessmentAttempt.passed == True, 1), else_=0)).label('passed_attempts')
        ).join(AssessmentAttempt, Batch.id == AssessmentAttempt.batch_id).filter(
            Batch.status == BatchStatus.ACTIVE
        ).group_by(Batch.id, Batch.name).all()
        
        batch_stats = []
        for batch_id, batch_name, total, passed in batch_performance:
            success_rate = round((passed / total * 100), 1) if total > 0 else 0
            batch_stats.append({
                "batch_id": str(batch_id),
                "batch_name": batch_name,
                "success_rate": success_rate,
                "total_assessments": total
            })
        
        batch_stats.sort(key=lambda x: x['success_rate'], reverse=True)

        # ========================================
        # 6. ASSESSMENT TRENDS (Last 7 days)
        # ========================================
        assessment_trend = []
        for i in range(7):
            day = utc_now() - timedelta(days=6-i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day.replace(hour=23, minute=59, second=59)

            day_attempts = db.query(func.count(AssessmentAttempt.id)).filter(
                and_(
                    AssessmentAttempt.evaluated_at >= day_start,
                    AssessmentAttempt.evaluated_at <= day_end
                )
            ).scalar() or 0

            day_passed = db.query(func.count(AssessmentAttempt.id)).filter(
                and_(
                    AssessmentAttempt.evaluated_at >= day_start,
                    AssessmentAttempt.evaluated_at <= day_end,
                    AssessmentAttempt.passed == True
                )
            ).scalar() or 0

            pass_rate = round((day_passed / day_attempts * 100), 1) if day_attempts > 0 else 0

            assessment_trend.append({
                "date": day.strftime("%a"),  # Mon, Tue, etc.
                "attempts": day_attempts,
                "pass_rate": pass_rate
            })

        # ========================================
        # 7. TRAINER EFFECTIVENESS
        # ========================================
        # Get trainer feedback ratings
        trainer_ratings = db.query(
            TrainerFeedback.trainer_id,
            User.name,
            func.avg((TrainerFeedback.subject_knowledge + TrainerFeedback.communication_skills +
                     TrainerFeedback.session_quality + TrainerFeedback.doubt_resolution) / 4.0).label('avg_rating'),
            func.count(TrainerFeedback.id).label('feedback_count')
        ).join(User, TrainerFeedback.trainer_id == User.id).group_by(
            TrainerFeedback.trainer_id, User.name
        ).all()

        trainer_stats = [
            {
                "trainer_id": str(trainer_id),
                "trainer_name": name,
                "avg_rating": round(float(avg_rating), 2),
                "feedback_count": count
            }
            for trainer_id, name, avg_rating, count in trainer_ratings
        ]

        # ========================================
        # 8. RESOURCE UTILIZATION
        # ========================================
        active_batches = db.query(Batch).filter(Batch.status == BatchStatus.ACTIVE).all()

        total_capacity = sum(batch.max_capacity or 0 for batch in active_batches)
        total_enrolled = sum(batch.current_enrollment or 0 for batch in active_batches)

        utilization_rate = round((total_enrolled / total_capacity * 100), 1) if total_capacity > 0 else 0

        # ========================================
        # 9. PROFILE REVIEW BACKLOG
        # ========================================
        pending_profiles = db.query(func.count(Maverick.id)).filter(
            Maverick.profile_status == ProfileStatus.PENDING
        ).scalar() or 0

        approved_last_7_days = db.query(func.count(Maverick.id)).filter(
            and_(
                Maverick.profile_status == ProfileStatus.APPROVED,
                Maverick.created_at >= utc_now() - timedelta(days=7)
            )
        ).scalar() or 0

        # ========================================
        # 10. DEPLOYMENT SUCCESS RATE
        # ========================================
        training_complete = db.query(func.count(func.distinct(MaverickJobProgress.maverick_id))).filter(
            MaverickJobProgress.status == JobProgressStatus.COMPLETED
        ).scalar() or 0

        actually_deployed = db.query(func.count(Maverick.id)).filter(
            Maverick.deployment_status == DeploymentStatus.DEPLOYED
        ).scalar() or 0

        deployment_success_rate = round((actually_deployed / training_complete * 100), 1) if training_complete > 0 else 0

        # ========================================
        # ADDITIONAL METRICS
        # ========================================
        total_mavericks = db.query(func.count(Maverick.id)).scalar() or 0
        active_batch_count = db.query(func.count(Batch.id)).filter(Batch.status == BatchStatus.ACTIVE).scalar() or 0

        pending_deployment_requests = db.query(func.count(DeploymentRequest.id)).filter(
            DeploymentRequest.status == DeploymentRequestStatus.PENDING
        ).scalar() or 0

        # Top 5 in-demand skills (from deployment requests and batch requirements)
        in_demand_skills = {}
        batches = db.query(Batch).filter(Batch.status == BatchStatus.ACTIVE).all()
        for batch in batches:
            if batch.required_skills:
                for skill in batch.required_skills:
                    in_demand_skills[skill] = in_demand_skills.get(skill, 0) + 1

        top_demand_skills = sorted(in_demand_skills.items(), key=lambda x: x[1], reverse=True)[:5]

        # ========================================
        # RETURN COMPREHENSIVE INSIGHTS
        # ========================================
        return {
            "summary": {
                "total_mavericks": total_mavericks,
                "active_batches": active_batch_count,
                "pending_reviews": pending_profiles,
                "deployment_requests": pending_deployment_requests
            },
            "insights": {
                # INSIGHT 1: Training Effectiveness
                "training_effectiveness": {
                    "success_rate": training_success_rate,
                    "total_attempts": total_assessment_attempts,
                    "passed_attempts": passed_attempts,
                    "failed_attempts": total_assessment_attempts - passed_attempts,
                    "status": "excellent" if training_success_rate >= 80 else "good" if training_success_rate >= 60 else "needs_improvement"
                },

                # INSIGHT 2: Deployment Pipeline Health
                "deployment_pipeline": {
                    "avg_days_to_deploy": avg_time_to_deploy,
                    "deployment_success_rate": deployment_success_rate,
                    "training_complete_count": training_complete,
                    "deployed_count": actually_deployed,
                    "status": "fast" if avg_time_to_deploy <= 90 else "moderate" if avg_time_to_deploy <= 120 else "slow"
                },

                # INSIGHT 3: Skill Gap Analysis
                "skill_gaps": {
                    "gaps": skill_gaps,
                    "total_skills_analyzed": len(skill_gaps),
                    "critical_gaps": [s for s in skill_gaps if s['avg_proficiency'] < 40]
                },

                # INSIGHT 4: At-Risk Mavericks
                "at_risk": {
                    "count": len(at_risk_list),
                    "mavericks": at_risk_list,
                    "requires_action": len(at_risk_list) > 0
                },

                # INSIGHT 5: Batch Performance
                "batch_performance": {
                    "batches": batch_stats[:5],  # Top 5
                    "best_batch": batch_stats[0] if batch_stats else None,
                    "worst_batch": batch_stats[-1] if batch_stats else None
                },

                # INSIGHT 6: Assessment Trends
                "assessment_trends": {
                    "weekly_data": assessment_trend,
                    "trend": "improving" if len(assessment_trend) > 1 and assessment_trend[-1]['pass_rate'] > assessment_trend[0]['pass_rate'] else "declining"
                },

                # INSIGHT 7: Trainer Effectiveness
                "trainer_effectiveness": {
                    "trainers": trainer_stats,
                    "avg_platform_rating": round(sum(t['avg_rating'] for t in trainer_stats) / len(trainer_stats), 2) if trainer_stats else 0
                },

                # INSIGHT 8: Resource Utilization
                "resource_utilization": {
                    "utilization_rate": utilization_rate,
                    "total_capacity": total_capacity,
                    "total_enrolled": total_enrolled,
                    "available_seats": total_capacity - total_enrolled,
                    "status": "optimal" if 70 <= utilization_rate <= 90 else "under_utilized" if utilization_rate < 70 else "over_capacity"
                },

                # INSIGHT 9: Profile Review Backlog
                "profile_backlog": {
                    "pending_count": pending_profiles,
                    "approved_last_week": approved_last_7_days,
                    "daily_review_rate": round(approved_last_7_days / 7, 1),
                    "days_to_clear": round(pending_profiles / (approved_last_7_days / 7)) if approved_last_7_days > 0 else 0
                },

                # INSIGHT 10: Top Demand Skills
                "in_demand_skills": [
                    {"skill": skill, "demand_count": count}
                    for skill, count in top_demand_skills
                ]
            },
            "period": {
                "days": days,
                "start_date": cutoff_date.isoformat(),
                "end_date": utc_now().isoformat()
            }
        }

    except Exception as e:
        logger.error(f"Analytics error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate analytics: {str(e)}"
        )


@router.get("/export/excel")
async def export_analytics_to_excel(
    days: int = 30,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_hr_user)
):
    """
    Export comprehensive HR analytics report to Excel

    Creates a professional Excel report with multiple sheets:
    - Summary Dashboard
    - Batch Performance
    - Skill Gaps
    - At-Risk Mavericks
    - Assessment Trends
    - Trainer Ratings
    """

    try:
        # Get all analytics data
        analytics_response = await get_analytics_overview(days, db, current_user)
        insights = analytics_response['insights']
        summary = analytics_response['summary']

        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Define styles
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=14, color="1e3a8a")

        # ========================================
        # SHEET 1: EXECUTIVE SUMMARY
        # ========================================
        ws_summary = wb.create_sheet("Executive Summary")
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        ws_summary.column_dimensions['C'].width = 40

        # Title
        ws_summary['A1'] = "HR ANALYTICS REPORT - MAVERICK ASCEND"
        ws_summary['A1'].font = Font(bold=True, size=16, color="1e3a8a")
        ws_summary.merge_cells('A1:C1')

        ws_summary['A2'] = f"Period: Last {days} days"
        ws_summary['A2'].font = Font(italic=True, size=10)

        # Key Metrics
        row = 4
        ws_summary[f'A{row}'] = "KEY METRICS"
        ws_summary[f'A{row}'].font = title_font
        row += 1

        metrics = [
            ("Total Mavericks", summary['total_mavericks'], ""),
            ("Active Batches", summary['active_batches'], ""),
            ("Pending Reviews", summary['pending_reviews'], "Action needed if > 10"),
            ("Training Success Rate", f"{insights['training_effectiveness']['success_rate']}%",
             "Excellent" if insights['training_effectiveness']['success_rate'] >= 80 else "Needs improvement"),
            ("Deployment Success Rate", f"{insights['deployment_pipeline']['deployment_success_rate']}%", ""),
            ("Avg Days to Deploy", insights['deployment_pipeline']['avg_days_to_deploy'],
             "Good" if insights['deployment_pipeline']['avg_days_to_deploy'] <= 90 else "Slow"),
            ("At-Risk Mavericks", insights['at_risk']['count'], "Requires intervention"),
            ("Resource Utilization", f"{insights['resource_utilization']['utilization_rate']}%",
             insights['resource_utilization']['status'].replace('_', ' ').title()),
        ]

        for metric_name, value, comment in metrics:
            ws_summary[f'A{row}'] = metric_name
            ws_summary[f'B{row}'] = value
            ws_summary[f'C{row}'] = comment
            ws_summary[f'A{row}'].font = Font(bold=True)
            row += 1

        # ========================================
        # SHEET 2: BATCH PERFORMANCE
        # ========================================
        ws_batch = wb.create_sheet("Batch Performance")
        ws_batch.column_dimensions['A'].width = 30
        ws_batch.column_dimensions['B'].width = 15
        ws_batch.column_dimensions['C'].width = 20

        ws_batch['A1'] = "BATCH PERFORMANCE ANALYSIS"
        ws_batch['A1'].font = title_font

        # Headers
        headers = ['Batch Name', 'Success Rate (%)', 'Total Assessments']
        for col, header in enumerate(headers, start=1):
            cell = ws_batch.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data
        row = 4
        for batch in insights['batch_performance']['batches']:
            ws_batch[f'A{row}'] = batch['batch_name']
            ws_batch[f'B{row}'] = batch['success_rate']
            ws_batch[f'C{row}'] = batch['total_assessments']

            # Color code success rate
            if batch['success_rate'] >= 80:
                ws_batch[f'B{row}'].fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
            elif batch['success_rate'] >= 60:
                ws_batch[f'B{row}'].fill = PatternFill(start_color="fef9c3", end_color="fef9c3", fill_type="solid")
            else:
                ws_batch[f'B{row}'].fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")

            row += 1

        # ========================================
        # SHEET 3: SKILL GAPS
        # ========================================
        ws_skills = wb.create_sheet("Skill Gaps")
        ws_skills.column_dimensions['A'].width = 25
        ws_skills.column_dimensions['B'].width = 20
        ws_skills.column_dimensions['C'].width = 15
        ws_skills.column_dimensions['D'].width = 30

        ws_skills['A1'] = "SKILL GAP ANALYSIS"
        ws_skills['A1'].font = title_font

        headers = ['Skill', 'Avg Proficiency', 'Maverick Count', 'Recommendation']
        for col, header in enumerate(headers, start=1):
            cell = ws_skills.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        row = 4
        for gap in insights['skill_gaps']['gaps']:
            ws_skills[f'A{row}'] = gap['skill']
            ws_skills[f'B{row}'] = gap['avg_proficiency']
            ws_skills[f'C{row}'] = gap['maverick_count']

            if gap['avg_proficiency'] < 40:
                ws_skills[f'D{row}'] = "🔴 CRITICAL - Immediate training needed"
                ws_skills[f'B{row}'].fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
            elif gap['avg_proficiency'] < 60:
                ws_skills[f'D{row}'] = "🟡 Schedule upskilling program"
                ws_skills[f'B{row}'].fill = PatternFill(start_color="fef9c3", end_color="fef9c3", fill_type="solid")

            row += 1

        # ========================================
        # SHEET 4: AT-RISK MAVERICKS
        # ========================================
        ws_risk = wb.create_sheet("At-Risk Mavericks")
        ws_risk.column_dimensions['A'].width = 30
        ws_risk.column_dimensions['B'].width = 20
        ws_risk.column_dimensions['C'].width = 40

        ws_risk['A1'] = "AT-RISK MAVERICKS - INTERVENTION REQUIRED"
        ws_risk['A1'].font = title_font

        headers = ['Maverick Name', 'Failed Attempts', 'Recommended Action']
        for col, header in enumerate(headers, start=1):
            cell = ws_risk.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        row = 4
        for maverick in insights['at_risk']['mavericks']:
            ws_risk[f'A{row}'] = maverick['name']
            ws_risk[f'B{row}'] = maverick['failed_attempts']

            if maverick['failed_attempts'] >= 3:
                ws_risk[f'C{row}'] = "Immediate mentoring + 1-on-1 support"
                ws_risk[f'B{row}'].fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
            else:
                ws_risk[f'C{row}'] = "Additional practice sessions"
                ws_risk[f'B{row}'].fill = PatternFill(start_color="fef9c3", end_color="fef9c3", fill_type="solid")

            row += 1

        # ========================================
        # SHEET 5: ASSESSMENT TRENDS
        # ========================================
        ws_trends = wb.create_sheet("Assessment Trends")
        ws_trends.column_dimensions['A'].width = 15
        ws_trends.column_dimensions['B'].width = 15
        ws_trends.column_dimensions['C'].width = 20

        ws_trends['A1'] = "7-DAY ASSESSMENT TRENDS"
        ws_trends['A1'].font = title_font

        headers = ['Day', 'Attempts', 'Pass Rate (%)']
        for col, header in enumerate(headers, start=1):
            cell = ws_trends.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        row = 4
        for trend in insights['assessment_trends']['weekly_data']:
            ws_trends[f'A{row}'] = trend['date']
            ws_trends[f'B{row}'] = trend['attempts']
            ws_trends[f'C{row}'] = trend['pass_rate']
            row += 1

        # ========================================
        # SHEET 6: TRAINER EFFECTIVENESS
        # ========================================
        ws_trainer = wb.create_sheet("Trainer Effectiveness")
        ws_trainer.column_dimensions['A'].width = 25
        ws_trainer.column_dimensions['B'].width = 15
        ws_trainer.column_dimensions['C'].width = 18

        ws_trainer['A1'] = "TRAINER EFFECTIVENESS RATINGS"
        ws_trainer['A1'].font = title_font

        headers = ['Trainer Name', 'Avg Rating', 'Feedback Count']
        for col, header in enumerate(headers, start=1):
            cell = ws_trainer.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        row = 4
        for trainer in insights['trainer_effectiveness']['trainers']:
            ws_trainer[f'A{row}'] = trainer['trainer_name']
            ws_trainer[f'B{row}'] = trainer['avg_rating']
            ws_trainer[f'C{row}'] = trainer['feedback_count']

            # Color code ratings
            if trainer['avg_rating'] >= 4.5:
                ws_trainer[f'B{row}'].fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
            elif trainer['avg_rating'] >= 3.5:
                ws_trainer[f'B{row}'].fill = PatternFill(start_color="fef9c3", end_color="fef9c3", fill_type="solid")
            else:
                ws_trainer[f'B{row}'].fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")

            row += 1

        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return as downloadable file
        filename = f"HR_Analytics_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logger.error(f"Excel export error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export analytics: {str(e)}"
        )
