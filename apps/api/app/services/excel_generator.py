"""
Excel Generator - Export Query Results to Excel
Generates formatted Excel files from query results with statistics
"""
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
import io

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    pd = None

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generate formatted Excel files from query results"""
    
    def __init__(self):
        if not EXCEL_AVAILABLE:
            logger.warning("Excel generation libraries not available. Install pandas and openpyxl.")
    
    def generate_excel(
        self,
        data: List[Dict[str, Any]],
        stats: Dict[str, Any],
        query_info: Optional[Dict[str, Any]] = None,
        filename_prefix: str = "query_results"
    ) -> io.BytesIO:
        """
        Generate Excel file from query results
        
        Args:
            data: List of row dictionaries from query results
            stats: Statistics dictionary from query executor
            query_info: Optional dict with natural_query, sql, explanation
            filename_prefix: Prefix for the generated filename
        
        Returns:
            BytesIO buffer containing Excel file
        
        Example:
            buffer = generator.generate_excel(data, stats, query_info)
            with open("results.xlsx", "wb") as f:
                f.write(buffer.getvalue())
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("pandas and openpyxl are required for Excel generation")
        
        logger.info(f"Generating Excel file with {len(data)} rows")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Sheet 1: Query Results
        self._create_results_sheet(wb, data, stats)
        
        # Sheet 2: Statistics
        self._create_statistics_sheet(wb, stats)
        
        # Sheet 3: Query Info (if provided)
        if query_info:
            self._create_query_info_sheet(wb, query_info, stats)
        
        # Save to BytesIO buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"✅ Excel file generated successfully")
        return buffer
    
    def _create_results_sheet(self, wb: Workbook, data: List[Dict[str, Any]], stats: Dict[str, Any]):
        """Create the main results sheet with data"""
        ws = wb.create_sheet("Query Results", 0)
        
        if not data:
            ws['A1'] = "No results found"
            return
        
        # Convert to DataFrame
        df = pd.DataFrame(data)

        # Write DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                # Convert lists/arrays to comma-separated strings for Excel
                if isinstance(value, (list, tuple)):
                    value = ', '.join(str(item) for item in value)
                elif isinstance(value, dict):
                    value = str(value)  # Convert dicts to string representation

                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        logger.info(f"   Results sheet created with {len(data)} rows, {len(df.columns)} columns")
    
    def _create_statistics_sheet(self, wb: Workbook, stats: Dict[str, Any]):
        """Create statistics sheet"""
        ws = wb.create_sheet("Statistics", 1)
        
        row = 1
        
        # Title
        ws['A1'] = "Query Statistics"
        ws['A1'].font = Font(bold=True, size=14)
        row += 2
        
        # Basic stats
        ws[f'A{row}'] = "Total Rows"
        ws[f'B{row}'] = stats.get('total_rows', 0)
        row += 1
        
        ws[f'A{row}'] = "Execution Time (ms)"
        ws[f'B{row}'] = stats.get('execution_time_ms', 0)
        row += 1
        
        ws[f'A{row}'] = "Columns"
        ws[f'B{row}'] = len(stats.get('columns', []))
        row += 2
        
        # Column list
        if stats.get('columns'):
            ws[f'A{row}'] = "Column Names"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            for col in stats['columns']:
                col_type = stats.get('column_types', {}).get(col, 'unknown')
                ws[f'A{row}'] = col
                ws[f'B{row}'] = col_type
                row += 1
            row += 1
        
        # Numeric aggregations
        aggregations = stats.get('aggregations', {})
        if aggregations.get('numeric'):
            ws[f'A{row}'] = "Numeric Column Statistics"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Header
            headers = ['Column', 'Count', 'Min', 'Max', 'Average', 'Sum']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            row += 1
            
            for col_name, agg_data in aggregations['numeric'].items():
                ws[f'A{row}'] = col_name
                ws[f'B{row}'] = agg_data.get('count', 0)
                ws[f'C{row}'] = agg_data.get('min', 0)
                ws[f'D{row}'] = agg_data.get('max', 0)
                ws[f'E{row}'] = agg_data.get('avg', 0)
                ws[f'F{row}'] = agg_data.get('sum', 0)
                row += 1
            row += 1
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2
        
        logger.info("   Statistics sheet created")
    
    def _create_query_info_sheet(self, wb: Workbook, query_info: Dict[str, Any], stats: Dict[str, Any]):
        """Create query information sheet"""
        ws = wb.create_sheet("Query Info", 2)
        
        row = 1
        
        # Title
        ws['A1'] = "Query Information"
        ws['A1'].font = Font(bold=True, size=14)
        row += 2
        
        # Natural language query
        if query_info.get('natural_query'):
            ws[f'A{row}'] = "Natural Language Query"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = query_info['natural_query']
            ws.merge_cells(f'A{row}:D{row}')
            row += 2
        
        # SQL Query
        if query_info.get('sql'):
            ws[f'A{row}'] = "Generated SQL"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = query_info['sql']
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 2
        
        # Explanation
        if query_info.get('explanation'):
            ws[f'A{row}'] = "Explanation"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = query_info['explanation']
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 2
        
        # Tables used
        if query_info.get('tables_used'):
            ws[f'A{row}'] = "Tables Used"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = ', '.join(query_info['tables_used'])
            row += 2
        
        # Execution info
        ws[f'A{row}'] = "Executed At"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = stats.get('executed_at', datetime.now().isoformat())
        row += 2
        
        ws[f'A{row}'] = "Total Rows Returned"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = stats.get('total_rows', 0)
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        
        logger.info("   Query info sheet created")


# Global instance
excel_generator = ExcelGenerator()
