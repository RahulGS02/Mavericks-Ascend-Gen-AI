"""
Test Excel Generation Only
Isolate and test just the Excel generation component
"""
import sys
from pathlib import Path

# Load environment
from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / '.env')

print("=" * 80)
print("🧪 Testing Excel Generation Specifically")
print("=" * 80)

# Test 1: Import Excel Generator
print("\n1️⃣  Importing excel_generator...")
try:
    from app.services.excel_generator import excel_generator
    print("✅ Excel generator imported")
except Exception as e:
    print(f"❌ Failed to import: {e}")
    sys.exit(1)

# Test 2: Check dependencies
print("\n2️⃣  Checking dependencies...")
try:
    import pandas as pd
    import openpyxl
    print(f"✅ pandas {pd.__version__}")
    print(f"✅ openpyxl {openpyxl.__version__}")
except ImportError as e:
    print(f"❌ Missing dependency: {e}")
    sys.exit(1)

# Test 3: Create sample data (similar to what query returns)
print("\n3️⃣  Creating sample data...")

sample_data = [
    {
        "id": "uuid-1",
        "name": "Alice Johnson",
        "email": "alice@example.com",
        "phone": "1234567890",
        "deployment_status": "AVAILABLE",
        "cgpa": 8.5
    },
    {
        "id": "uuid-2",
        "name": "Bob Smith",
        "email": "bob@example.com",
        "phone": "0987654321",
        "deployment_status": "DEPLOYED",
        "cgpa": 9.2
    },
    {
        "id": "uuid-3",
        "name": "Charlie Brown",
        "email": "charlie@example.com",
        "phone": "5555555555",
        "deployment_status": "AVAILABLE",
        "cgpa": 7.8
    }
]

sample_stats = {
    "total_rows": 3,
    "columns": ["id", "name", "email", "phone", "deployment_status", "cgpa"],
    "column_types": {
        "id": "uuid",
        "name": "string",
        "email": "string",
        "phone": "string",
        "deployment_status": "string",
        "cgpa": "float"
    },
    "execution_time_ms": 123.45,
    "executed_at": "2026-05-23T10:30:00",
    "aggregations": {
        "numeric": {
            "cgpa": {
                "count": 3,
                "min": 7.8,
                "max": 9.2,
                "avg": 8.5,
                "sum": 25.5
            }
        },
        "string": {
            "deployment_status": {
                "count": 3,
                "unique_count": 2,
                "unique_values": ["AVAILABLE", "DEPLOYED"]
            }
        }
    }
}

sample_query_info = {
    "natural_query": "Show me available mavericks",
    "sql": "SELECT * FROM mavericks WHERE deployment_status = 'AVAILABLE' LIMIT 100",
    "explanation": "Returns all mavericks with AVAILABLE deployment status",
    "tables_used": ["mavericks"]
}

print(f"✅ Sample data created: {len(sample_data)} rows")

# Test 4: Generate Excel
print("\n4️⃣  Generating Excel file...")
try:
    buffer = excel_generator.generate_excel(sample_data, sample_stats, sample_query_info)
    print(f"✅ Excel buffer created")
    print(f"   Buffer position: {buffer.tell()}")
    print(f"   Buffer type: {type(buffer)}")
    
    # Check buffer size
    buffer.seek(0, 2)  # Seek to end
    size = buffer.tell()
    buffer.seek(0)  # Seek back to start
    
    print(f"   Buffer size: {size} bytes")
    
    if size == 0:
        print("❌ ERROR: Buffer is empty!")
        sys.exit(1)
    
except Exception as e:
    print(f"❌ Excel generation failed: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# Test 5: Save to file
print("\n5️⃣  Saving Excel file...")
try:
    filename = "test_excel_diagnostic.xlsx"
    with open(filename, "wb") as f:
        f.write(buffer.getvalue())
    
    file_size = Path(filename).stat().st_size
    print(f"✅ File saved: {filename}")
    print(f"   File size: {file_size} bytes")
    
    if file_size == 0:
        print("❌ ERROR: File is empty!")
        sys.exit(1)
    
except Exception as e:
    print(f"❌ Failed to save file: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# Test 6: Try to read it back
print("\n6️⃣  Verifying Excel file...")
try:
    import openpyxl
    wb = openpyxl.load_workbook(filename)
    print(f"✅ Excel file is valid")
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   Number of sheets: {len(wb.sheetnames)}")
    
    # Check first sheet
    ws = wb[wb.sheetnames[0]]
    row_count = ws.max_row
    col_count = ws.max_column
    print(f"   First sheet rows: {row_count}")
    print(f"   First sheet columns: {col_count}")
    
except Exception as e:
    print(f"❌ Failed to read Excel file: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print("\n" + "=" * 80)
print("✅ ALL EXCEL TESTS PASSED!")
print("=" * 80)
print("\n💡 Excel generation is working correctly!")
print(f"   Generated file: {filename}")
print(f"   Size: {file_size} bytes")
print(f"   Sheets: {len(wb.sheetnames)}")
print("\n🔍 This means the issue is likely in the API endpoint or streaming.")
print("   Check server logs for the actual error when downloading.")
